"""文件摄取：上传 ≠ 长期记忆。显式「存入记忆」才切片入库（P0-02 §5.5）。"""

from __future__ import annotations

import io
from uuid import UUID

from cogniwork.core.errors import InvalidRequest

from .models import MemoryStatus, MemoryType, SourceType
from .service import MemoryService

MAX_INGEST_BYTES = 20 * 1024 * 1024
MAX_PAGES = 200
CHUNK_CHARS = 2000  # ~500 tokens
OVERLAP_CHARS = 320  # ~80 tokens


def ingest_file(
    service: MemoryService,
    user_id: UUID,
    *,
    filename: str,
    content: bytes,
    file_id: str | None = None,
) -> list:
    if len(content) > MAX_INGEST_BYTES:
        raise InvalidRequest(
            "This file is larger than 20MB. Split it before saving to memory.",
            details={"max_bytes": MAX_INGEST_BYTES},
        )
    text = extract_text(filename, content)
    chunks = chunk_text(text)
    if not chunks:
        raise InvalidRequest("I could not find any text in that file.")
    items = []
    for index, chunk in enumerate(chunks):
        items.append(
            service.create(
                user_id,
                type=MemoryType.SEMANTIC,
                content=chunk,
                summary=chunk.splitlines()[0][:120] if chunk.strip() else filename,
                subtype="file_chunk",
                importance=3,
                source_type=SourceType.FILE_INGEST,
                source_ref={
                    "file_id": file_id,
                    "filename": filename,
                    "page": index + 1,
                    "quote": chunk[:180],
                },
                status=MemoryStatus.PENDING,
            )
        )
    return items


def extract_text(filename: str, content: bytes) -> str:
    lower = filename.lower()
    if lower.endswith(".pdf"):
        return _pdf(content)
    if lower.endswith(".docx"):
        return _docx(content)
    if lower.endswith(".xlsx"):
        return _xlsx(content)
    try:
        return content.decode("utf-8")
    except UnicodeDecodeError:
        return content.decode("utf-8", errors="replace")


def chunk_text(text: str) -> list[str]:
    blocks = _split_blocks(text)
    chunks: list[str] = []
    buf = ""
    for block in blocks:
        heading = block.lstrip().startswith("#")
        candidate = f"{buf}\n\n{block}".strip() if buf else block
        if heading and buf:
            chunks.append(buf)
            buf = block
            continue
        if len(candidate) <= CHUNK_CHARS:
            buf = candidate
            continue
        if buf:
            chunks.append(buf)
        if len(block) <= CHUNK_CHARS:
            overlap = buf[-OVERLAP_CHARS:] if buf else ""
            buf = f"{overlap}\n\n{block}".strip() if overlap else block
        else:
            for start in range(0, len(block), CHUNK_CHARS - OVERLAP_CHARS):
                chunks.append(block[start : start + CHUNK_CHARS].strip())
            buf = ""
    if buf:
        chunks.append(buf)
    return [chunk for chunk in chunks if chunk.strip()]


def _split_blocks(text: str) -> list[str]:
    parts: list[str] = []
    buf: list[str] = []
    for line in text.splitlines():
        if line.startswith("#") and buf:
            parts.append("\n".join(buf).strip())
            buf = [line]
        elif not line.strip() and buf:
            parts.append("\n".join(buf).strip())
            buf = []
        else:
            buf.append(line)
    if buf:
        parts.append("\n".join(buf).strip())
    return [part for part in parts if part]


def _pdf(content: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise InvalidRequest("PDF support is not installed.") from exc
    reader = PdfReader(io.BytesIO(content))
    if len(reader.pages) > MAX_PAGES:
        raise InvalidRequest(
            "This PDF has more than 200 pages. Split it before saving to memory.",
            details={"max_pages": MAX_PAGES},
        )
    pages = [(page.extract_text() or "") for page in reader.pages]
    return "\n\n".join(pages)


def _docx(content: bytes) -> str:
    try:
        from docx import Document
    except ImportError as exc:
        raise InvalidRequest("DOCX support is not installed.") from exc
    document = Document(io.BytesIO(content))
    return "\n".join(p.text for p in document.paragraphs if p.text)


def _xlsx(content: bytes) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), data_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"# {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = ["" if cell is None else str(cell) for cell in row]
            if any(cell.strip() for cell in cells):
                lines.append(" | ".join(cells))
    return "\n".join(lines)
