"""产物预览（P0-04 WS-5：xlsx / csv / md / docx / png）。

解析放在后端，不放浏览器。三个理由，按重要性排：

1. 解析器已经在这边了（`memory/ingest.py` 用的 openpyxl / python-docx）。
   浏览器侧要重新引一套 xlsx 与 docx 解析器，等于同一件事维护两份。
2. 产物下载要带 Bearer token，`<img src>` / `<iframe>` 带不了 header。
   png 走这里转成 data URI 才显示得出来。
3. 预览是**截断**的。截断在服务端做，才不用把一个 20MB 的 xlsx 整个送到浏览器
   才决定只显示前 50 行。

预览不改任何东西，也不额外采集 —— 它读的就是用户自己刚拿到的那份产物，
所以不需要 Scope（`00-conventions.md` §3 注同理）。
"""

from __future__ import annotations

import base64
import csv
import io
from typing import Any

MAX_ROWS = 50
MAX_COLS = 20
MAX_TEXT_CHARS = 20_000
# 图片走 data URI，整个 JSON 都会被浏览器 hold 在内存里，给个上限。
MAX_IMAGE_BYTES = 2 * 1024 * 1024

TABLE_SUFFIXES = (".xlsx", ".csv")
TEXT_SUFFIXES = (".md", ".txt", ".json", ".docx")
IMAGE_SUFFIXES = (".png", ".jpg", ".jpeg", ".gif", ".webp")


def build_preview(filename: str, content_type: str, content: bytes) -> dict[str, Any]:
    """收成一个前端只需要认四种 kind 的结构。

    认不出的格式返回 `kind="none"` 而不是报错 —— 「这个格式还不能预览，下载吧」
    是一个正常状态，不是一次失败。
    """
    lower = (filename or "").lower()
    if lower.endswith(".xlsx"):
        return _xlsx(content)
    if lower.endswith(".csv"):
        return _csv(content)
    if lower.endswith(".docx"):
        return _text(_docx_text(content), kind="text")
    if lower.endswith(".md"):
        return _text(_decode(content), kind="markdown")
    if lower.endswith((".txt", ".json")):
        return _text(_decode(content), kind="text")
    if lower.endswith(IMAGE_SUFFIXES):
        return _image(content_type or "image/png", content)
    return {"kind": "none", "reason": "unsupported_format"}


def _xlsx(content: bytes) -> dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.worksheets[0] if workbook.worksheets else None
    if sheet is None:
        return {"kind": "none", "reason": "empty_workbook"}
    rows: list[list[str]] = []
    total = 0
    for row in sheet.iter_rows(values_only=True):
        total += 1
        if len(rows) < MAX_ROWS:
            cells = ["" if cell is None else str(cell) for cell in row[:MAX_COLS]]
            if any(cell.strip() for cell in cells):
                rows.append(cells)
    workbook.close()
    return _table(rows, total, sheet_name=str(sheet.title))


def _csv(content: bytes) -> dict[str, Any]:
    text = _decode(content)
    reader = csv.reader(io.StringIO(text))
    rows: list[list[str]] = []
    total = 0
    for row in reader:
        total += 1
        if len(rows) < MAX_ROWS:
            rows.append([str(cell) for cell in row[:MAX_COLS]])
    return _table(rows, total)


def _table(rows: list[list[str]], total: int, sheet_name: str | None = None) -> dict[str, Any]:
    header = rows[0] if rows else []
    return {
        "kind": "table",
        "sheet_name": sheet_name,
        "header": header,
        "rows": rows[1:],
        "row_count": max(total - 1, 0),
        "truncated": total > MAX_ROWS,
    }


def _text(text: str, *, kind: str) -> dict[str, Any]:
    clipped = text[:MAX_TEXT_CHARS]
    return {"kind": kind, "text": clipped, "truncated": len(text) > len(clipped)}


def _image(content_type: str, content: bytes) -> dict[str, Any]:
    if len(content) > MAX_IMAGE_BYTES:
        return {"kind": "none", "reason": "image_too_large"}
    encoded = base64.b64encode(content).decode()
    return {"kind": "image", "data_uri": f"data:{content_type};base64,{encoded}"}


def _docx_text(content: bytes) -> str:
    from cogniwork.memory.ingest import extract_text

    return extract_text("preview.docx", content)


def _decode(content: bytes) -> str:
    try:
        return content.decode("utf-8")
    except UnicodeDecodeError:
        return content.decode("utf-8", errors="replace")
