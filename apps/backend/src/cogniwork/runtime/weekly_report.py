"""把表格整理成周报 —— 零授权核心路径的确定性实现。

有 LLM key 时仍走工具调用；没有 key 时 stub 模型调用同一份逻辑。
这样 E2E 不依赖外网，演示在无密钥环境也能拿到真实产物。
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import UTC, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


@dataclass(frozen=True, slots=True)
class SheetTable:
    headers: list[str]
    rows: list[list[str]]


def parse_tabular(content: bytes, filename: str) -> SheetTable:
    lower = filename.lower()
    if lower.endswith(".csv") or lower.endswith(".txt"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        data = [list(r) for r in reader if any(cell.strip() for cell in r)]
        if not data:
            return SheetTable([], [])
        return SheetTable([str(c) for c in data[0]], [[str(c) for c in r] for r in data[1:]])

    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell) for cell in row]
            if any(v.strip() for v in values):
                data.append(values)
    finally:
        workbook.close()
    if not data:
        return SheetTable([], [])
    width = max(len(r) for r in data)
    normalized = [r + [""] * (width - len(r)) for r in data]
    return SheetTable(normalized[0], normalized[1:])


def _is_number(value: str) -> bool:
    try:
        float(value.replace(",", "").replace("%", "").strip())
        return bool(value.strip())
    except ValueError:
        return False


def _to_number(value: str) -> float:
    return float(value.replace(",", "").replace("%", "").strip())


def build_weekly_report(content: bytes, filename: str) -> tuple[str, bytes]:
    """返回 (markdown, xlsx bytes)。"""
    table = parse_tabular(content, filename)
    generated = datetime.now(UTC).strftime("%Y-%m-%d")
    if not table.headers:
        markdown = (
            f"# Weekly report\n\nGenerated {generated} from `{filename}`.\n\n"
            "The uploaded file had no tabular rows I could read.\n"
        )
        return markdown, _xlsx_from_markdown_fallback(markdown)

    numeric_cols: list[int] = []
    for idx in range(len(table.headers)):
        values = [row[idx] if idx < len(row) else "" for row in table.rows]
        nonempty = [v for v in values if v.strip()]
        if nonempty and all(_is_number(v) for v in nonempty):
            numeric_cols.append(idx)

    totals: dict[str, float] = {}
    for idx in numeric_cols:
        totals[table.headers[idx]] = sum(
            _to_number(row[idx]) for row in table.rows if idx < len(row) and row[idx].strip()
        )

    wow_lines: list[str] = []
    if len(numeric_cols) >= 2:
        current_idx, previous_idx = numeric_cols[0], numeric_cols[1]
        for row in table.rows:
            label = row[0] if row else ""
            if len(row) <= max(current_idx, previous_idx):
                continue
            if not row[current_idx].strip() or not row[previous_idx].strip():
                continue
            current = _to_number(row[current_idx])
            previous = _to_number(row[previous_idx])
            delta = current - previous
            pct = (delta / previous * 100) if previous else 0.0
            wow_lines.append(
                f"- **{label}**: {current:g} vs {previous:g} ({delta:+.1f}, {pct:+.1f}%)"
            )

    conclusions = _conclusions(table, numeric_cols, totals)
    header_line = " | ".join(table.headers)
    sep_line = " | ".join("---" for _ in table.headers)
    body_lines = [" | ".join(row[: len(table.headers)]) for row in table.rows[:50]]

    markdown = "\n".join(
        [
            "# Weekly report",
            "",
            f"Generated {generated} from `{filename}`.",
            "",
            "## Summary table",
            "",
            header_line,
            sep_line,
            *body_lines,
            "",
            "## Totals",
            "",
            *(
                [f"- **{name}**: {value:g}" for name, value in totals.items()]
                or ["- No numeric columns detected."]
            ),
            "",
            "## Week-over-week",
            "",
            *(wow_lines or ["- Not enough numeric columns to compute a comparison."]),
            "",
            "## Three takeaways",
            "",
            *conclusions,
            "",
        ]
    )
    return markdown, _xlsx_report(table, totals, conclusions, generated, filename)


def _conclusions(table: SheetTable, numeric_cols: list[int], totals: dict[str, float]) -> list[str]:
    if not table.rows:
        return [
            "1. The sheet is empty — nothing to summarize this week.",
            "2. Re-upload a sheet with headers and at least one data row.",
            "3. Keep channel names in the first column so comparisons stay readable.",
        ]
    label = table.headers[0] if table.headers else "row"
    if numeric_cols:
        idx = numeric_cols[0]
        ranked: list[tuple[str, float]] = []
        for row in table.rows:
            if idx < len(row) and row[idx].strip() and _is_number(row[idx]):
                ranked.append((row[0] if row else "", _to_number(row[idx])))
        ranked.sort(key=lambda item: item[1], reverse=True)
        top = ranked[0] if ranked else ("n/a", 0.0)
        bottom = ranked[-1] if ranked else ("n/a", 0.0)
        total_name, total_value = next(iter(totals.items()), ("metric", 0.0))
        return [
            f"1. **{top[0] or label}** led this week at {top[1]:g} on {table.headers[idx]}.",
            f"2. Combined {total_name} across {len(table.rows)} "
            f"{label.lower()}s is {total_value:g}.",
            f"3. **{bottom[0] or label}** is the lagging line ({bottom[1]:g}) "
            "— worth a look next week.",
        ]
    return [
        f"1. {len(table.rows)} {label.lower()}s were listed; none of the columns looked numeric.",
        "2. The table is included as-is so you can still share a formatted copy.",
        "3. Add a numeric column (this week / last week) next time for automatic deltas.",
    ]


def _xlsx_report(
    table: SheetTable,
    totals: dict[str, float],
    conclusions: list[str],
    generated: str,
    filename: str,
) -> bytes:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Weekly report"
    sheet["A1"] = "Weekly report"
    sheet["A1"].font = Font(bold=True, size=16)
    sheet["A2"] = f"Generated {generated} from {filename}"
    start = 4
    for col, header in enumerate(table.headers, start=1):
        cell = sheet.cell(start, col, header)
        cell.font = Font(bold=True)
    for r_idx, row in enumerate(table.rows, start=start + 1):
        for c_idx, value in enumerate(row[: len(table.headers)], start=1):
            parsed: str | float = value
            if _is_number(value):
                parsed = _to_number(value)
            sheet.cell(r_idx, c_idx, parsed)
    totals_row = start + 1 + len(table.rows) + 2
    sheet.cell(totals_row, 1, "Totals").font = Font(bold=True)
    for offset, (name, value) in enumerate(totals.items()):
        sheet.cell(totals_row + 1 + offset, 1, name)
        sheet.cell(totals_row + 1 + offset, 2, value)
    takeaway_row = totals_row + 1 + max(len(totals), 1) + 2
    sheet.cell(takeaway_row, 1, "Three takeaways").font = Font(bold=True)
    for offset, line in enumerate(conclusions):
        sheet.cell(takeaway_row + 1 + offset, 1, line)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _xlsx_from_markdown_fallback(markdown: str) -> bytes:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Weekly report"
    for idx, line in enumerate(markdown.splitlines(), start=1):
        sheet.cell(idx, 1, line)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def table_as_text(content: bytes, filename: str) -> str:
    table = parse_tabular(content, filename)
    if not table.headers:
        return f"(empty table from {filename})"
    lines = [", ".join(table.headers)]
    for row in table.rows:
        lines.append(", ".join(row[: len(table.headers)]))
    return "\n".join(lines)
